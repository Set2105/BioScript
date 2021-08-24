import openpyxl
from openpyxl.utils import get_column_letter
import os
import json
import re


def get_cell(coord: list):
    return get_column_letter(coord[1]) + str(coord[0])


class Cell:
    row = 0
    cell = 0

    def __init__(self, row: int = 0, cell: int = 0):
        self.row = row
        self.cell = cell

    def __call__(self, a: int = 0, b: int = 0):
        return get_column_letter(self.cell+b) + str(self.row + a)

    def __repr__(self):
        return str(get_column_letter(self.cell) + str(self.row))

    def add(self, a: int, b: int):
        self.row += a
        self.cell += b


class XLSXParser:
    source_files_path = ''
    result_file_path = ''
    key_words_dict_path = ''
    values_stack_path = ''

    values_stack = ''

    value_name_dict = {}
    value_name_stack = []

    pure_data = []
    sorted_data = {}
    key_words_dict = []

    start_cell = Cell(1, 5)
    cell_interval = 13

    wb = openpyxl.Workbook()
    ws = ''

    def __init__(self,
                 source_files_path='options',
                 result_file_path='result.xlsx',
                 key_words_dict_path='options/key_words_dict.json',
                 values_stack_path='options/gen_name.json'):

        self.source_files_path = source_files_path
        self.result_file_path = result_file_path
        self.key_words_dict_path = key_words_dict_path
        self.values_stack_path = values_stack_path

        try:
            with open(self.key_words_dict_path, 'r') as json_file:
                self.key_words_dict = json.load(json_file)
        except Exception as e:
            print(f'\033[31mFailed load KeyWordsDict json from \'{self.key_words_dict_path}\'\n {e.args}\033[0m')

        try:
            with open(self.values_stack_path, 'r') as json_file:
                self.value_name_dict = json.load(json_file)
        except Exception as e:
            print(f'\033[31mFailed load KeyWordsDict json from\'{self.values_stack_path}\'\n {e.args}\033[0m')

    def get_sheet_name(self, file_name):
        for key in self.value_name_dict.keys():
            if re.search(self.value_name_dict[key][0], file_name):
                return key
        return 'sheet'

    def read(self, file_name):
        self.wb = openpyxl.load_workbook(filename=file_name, read_only=True)
        self.ws = self.wb.worksheets[0]

        for row in self.ws.values:
            print('')
            for col in row:
                print(col, end=' ')

    def load_data(self, file_path: str):
        self.wb = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
        self.ws = self.wb.worksheets[0]
        for key_list in self.key_words_dict:
            parse_data = False
            key_words = key_list['key_words']
            for row in self.ws.values:
                if parse_data:
                    data_col = {}
                    data_col.update({'file_name': re.split('/', file_path)[-1]})
                    for i in range(0, len(key_words)):
                        if key_words[i] in key_list['replace_keys'].keys():
                            data_col.update({key_list['replace_keys'][key_words[i]]: row[i]})
                    if data_col['sort_val'] in key_list['sort_values']:
                        data_col.pop('sort_val')
                        self.pure_data.append(data_col)
                else:
                    if key_words:
                        if row[0] == key_words[0]:
                            parse_data = True
                            for i in range(1, len(key_words)):
                                if row[i] != key_words[i]:
                                    parse_data = False
                                    break

    def load_files(self, dir_path: str = source_files_path):
        file_path_list = list(map(lambda file_name: dir_path + '/' + file_name, os.listdir(dir_path)))
        for file_path in file_path_list:
            try:
                self.load_data(file_path)
                print(f'Successfully loaded data from \'{file_path}\'')
            except Exception as e:
                print(f'\033[31mFailed load \'{file_path}\'\n {e.args}\033[0m')

    def sort_data(self):
        for col in self.pure_data:
            if col['value'] is None:
                col['value'] = ''
            if col['file_name'] not in self.sorted_data.keys():
                self.sorted_data.update({col['file_name']: {col['instance']: {col['variable_name']: [col['value']]}}})
            else:
                if col['instance'] not in self.sorted_data[col['file_name']].keys():
                    self.sorted_data[col['file_name']].update({col['instance']: {col['variable_name']: [col['value']]}})
                else:
                    if col['variable_name'] not in self.sorted_data[col['file_name']][col['instance']].keys():
                        self.sorted_data[col['file_name']][col['instance']].update({col['variable_name']: [col['value']]})
                    else:
                        self.sorted_data[col['file_name']][col['instance']][col['variable_name']].append(col['value'])

    def print_pure_data(self):
        for item in self.pure_data:
            print(item)

    def print_sorted_data(self):
        for key1 in self.sorted_data:
            print(key1)
            for key2 in self.sorted_data[key1]:
                print(' ', key2)
                for key3 in self.sorted_data[key1][key2]:
                    print('  ', key3, self.sorted_data[key1][key2][key3])

    def initialize_value_name_stack(self, sheet_name):
        for key in self.value_name_dict.keys():
            if sheet_name == key:
                self.value_name_stack = self.value_name_dict[key]

    def initialize_write_sheet(self, sheet_name: str):
        self.ws = self.wb[sheet_name]

    def format_data_table(self, file_name: str, coord: Cell):
        self.ws.merge_cells(f'{coord(-1, 0)}:{coord(0, self.cell_interval-1)}')
        self.ws[coord(-1, 0)] = file_name

        coord.add(1, 0)
        self.ws.merge_cells(f'{coord()}:{coord(1, 0)}')
        self.ws[coord()] = 'Sample'

        coord.add(0, 1)
        for gen_name in self.value_name_stack:
            self.ws.merge_cells(f'{str(coord)}:{coord(0, 1)}')
            self.ws[str(coord)].value = gen_name
            self.ws[coord(1, 0)].value = 'Ct'
            self.ws[coord(1, 1)].value = 'Ct mean'
            coord.add(0, 2)

        for i in range(2):
            self.ws[coord()].value = self.value_name_stack[0]
            coord.add(0, 1)
            self.ws[coord()].value = self.value_name_stack[1]
            coord.add(0, 1)

    def find_free_space(self, sheet_name: str):
        self.ws = self.wb[sheet_name]
        verifiable_cell = Cell(self.start_cell.row, self.start_cell.cell)
        while True:
            if self.ws[verifiable_cell(0, 0)].value is None:
                break
            verifiable_cell.add(0, self.cell_interval)
        verifiable_cell.add(1, 0)
        return verifiable_cell

    def write_data_table(self, data: dict, coord: Cell, file_name):
        calibrator_name = input(f'{file_name}\nInput Calibrator value from {data.keys()}:')
        for key in data.keys():
            if key == calibrator_name:
                self.ws.merge_cells(f'{coord(1, 0)}:{coord(3, 0)}')
                self.ws[coord(1, 0)] = key
                self.write_data_row(data[key], Cell(coord.row, coord.cell+1))
        data.pop(calibrator_name)
        coord.add(3, 0)
        for key in data.keys():
            self.ws.merge_cells(f'{coord(1, 0)}:{coord(3, 0)}')
            self.ws[coord(1, 0)] = key
            self.write_data_row(data[key], Cell(coord.row, coord.cell+1))
            coord.add(3, 0)

    def write_data_row(self, data: dict, coord: Cell):
        for gen_name in self.value_name_stack:
            try:
                self.write_cell(data[gen_name], coord)
                coord.add(0, 2)
            except KeyError as e:
                print(f'Key error: {e.args} ')

    def write_cell(self, value_list: list, coord: Cell):
        i = 0
        for value in value_list:
            i += 1
            self.ws[coord(i, 0)] = value

    def write_xlsx(self):
        try:
            self.wb = openpyxl.load_workbook(self.result_file_path)
        except Exception as e:
            print(e)
            self.wb = openpyxl.Workbook(write_only=False)

        for file_name in self.sorted_data.keys():
            sheet_name = self.get_sheet_name(file_name)
            try:
                self.wb[sheet_name]
            except Exception as e:
                print(e.args)
                self.wb.create_sheet(sheet_name)
                print(f'sheet \'{sheet_name}\' was created')
            self.initialize_write_sheet(sheet_name)
            self.initialize_value_name_stack(sheet_name)
            written_coord = self.find_free_space(sheet_name)
            self.format_data_table(file_name, Cell(written_coord.row, written_coord.cell))
            written_coord.add(2, 0)
            self.write_data_table(self.sorted_data[file_name], Cell(written_coord.row, written_coord.cell), file_name)

        self.wb.save(self.result_file_path)
        print(f'WorkBook \'{self.result_file_path}\' was saved')


parser = XLSXParser()
parser.load_files('source files')
parser.sort_data()
parser.write_xlsx()
